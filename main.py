import collections
import threading
import time
import random
import openpyxl
from tkinter import *
from tkinter import ttk


class App:
    def __init__(self):
        self.STATE1 = "：正在送出礼物"
        self.STATE2 = "（已获得）"
        self.root = Tk()
        self.running_flag = False   # 设置滚动 flag
        self.time_span = 0.03       # 名称显示间隔
        self.qunyou_list, self.tianxuan_count, self.last_year_safufu = self.get_people_list()   # 拿到群友名单
        self.qunyou_temp = self.qunyou_list + [" "]
        self.safufu_temp = []
        for qunyou in self.qunyou_list:
            if qunyou == self.last_year_safufu:
                self.safufu_temp.append(self.last_year_safufu + "Safufu")
            else:
                self.safufu_temp.append(qunyou)
        self.qunyou_gift_remain = len(self.qunyou_list)
        self.tianxuan_gift_remain = self.tianxuan_count
        self.safufu_gift_remain = 1
        self.result_message_list = [""] * (len(self.qunyou_list) + self.tianxuan_count + 1)     # 这是结果消息显示列表
        self.result_gift_index = 0                                                              # 这是结果消息显示的 index
        self.ready_message_dict = dict()
        for qunyou in self.qunyou_temp:
            self.ready_message_dict[qunyou] = [""]
        self.ready_message_list = [""] * (len(self.qunyou_list) + 1)    # 这是准备消息显示列表
        self.get_ready_gift_index = 0                                   # 这是准备抽取礼物的 index
        self.get_gift_list = collections.deque()                        # 这是一个临时数组，存储已获得礼物的人 + 正在抽取礼物的人
        self.ready_message_dict[self.qunyou_temp[self.get_ready_gift_index]].insert(0, self.STATE1)
        self.get_gift_list.appendleft(self.qunyou_list[self.get_ready_gift_index])
        self.root.title("致远星联络中心专用抽奖 v1.0")
        width = 600
        height = 550
        left = (self.root.winfo_screenwidth() - width) / 2
        top = (self.root.winfo_screenheight() - height) / 2
        self.root.geometry("%dx%d+%d+%d" % (width, height, left, top))
        self.root.resizable(0, 0)
        self.create_widget()        # 创建组件
        self.set_widget()           # 初始化组件
        self.place_widget()         # 放置组件
        self.root.mainloop()

    def function_random(self, people):
        random.seed(time.time())
        return random.choice(people)

    def get_people_list(self):
        workbook = openpyxl.load_workbook("./peopleList.xlsx", data_only=True)
        res = []
        sheet = workbook[workbook.sheetnames[0]]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == "":
                continue
            else:
                res.append(sheet.cell(row=row, column=1).value)
        extra_gift = sheet.cell(row=2, column=2).value
        safufu = sheet.cell(row=2, column=3).value
        if isinstance(extra_gift, int):
            pass
        else:
            extra_gift = 1
        return res, extra_gift, str(safufu)

    def create_widget(self):
        # 名字显示组件 —— Label
        self.label_show_name_var = StringVar()
        self.label_show_name = ttk.Label(self.root,
                                         textvariable=self.label_show_name_var,
                                         font=('Arial', 50, 'bold'))

        # 开始按钮组件 —— Button
        self.start_button = ttk.Button(self.root,
                                       text="开始")

        # 抽奖方式组件 —— LabelFrame
        self.label_show_type_label = ttk.LabelFrame(self.root,
                                                    text="")

        # 抽奖种类单选按钮组件 —— Radiobutton
        # 只有随机抽奖才能触发 self.label_show_gift_label 选择奖品种类
        self.draw_type_button_var = IntVar()
        self.draw_type_button_var.set(1)
        self.draw_type_exchange = ttk.Radiobutton(self.label_show_type_label,
                                                  text="礼物交换",
                                                  variable=self.draw_type_button_var,
                                                  value=1,
                                                  command=lambda: self.hidden_or_show_gift_type())
        self.draw_type_random = ttk.Radiobutton(self.label_show_type_label,
                                                text="随机抽奖",
                                                variable=self.draw_type_button_var,
                                                value=2,
                                                command=lambda: self.hidden_or_show_gift_type())

        # 随机抽奖中，奖品种类组件 —— LabelFrame
        self.label_show_gift_label = ttk.LabelFrame(self.root,
                                                    text="")

        # 奖品种类单选按钮组件 —— Radiobutton
        self.gift_type_button_var = IntVar()
        self.gift_type_tianxuan = ttk.Radiobutton(self.label_show_gift_label,
                                                  text="天选之人",
                                                  variable=self.gift_type_button_var,
                                                  value=1,
                                                  state=DISABLED,
                                                  command=lambda: self.start_button_awake())
        self.gift_type_safufu = ttk.Radiobutton(self.label_show_gift_label,
                                                text="Safufu",
                                                variable=self.gift_type_button_var,
                                                value=2,
                                                state=DISABLED,
                                                command=lambda: self.start_button_awake())

        # 名单/概率显示组件 —— Message
        self.remain_message_var = StringVar()
        self.remain_message = Message(textvariable=self.remain_message_var,
                                      bg='#ffffff',
                                      relief=SUNKEN,
                                      anchor='nw')

        # 目前抽取的礼物显示组件 —— Label
        self.remain_label_var = StringVar()
        self.remain_label = ttk.Label

        # 结果显示组件 —— Message
        self.result_message_var = StringVar()
        self.result_message = Message(textvariable=self.result_message_var,
                                      bg='#ffffff',
                                      relief=SUNKEN,
                                      anchor='nw',
                                      justify=LEFT)

    def set_widget(self):
        default_name_ = '致远星'
        self.label_show_name_var.set(default_name_)
        self.label_show_name_adjust(default_name_)
        self.start_button.config(command=lambda: self.thread_it(self.get_draw_type))
        self.remain_message_var.set(self.get_ready_message())

    def place_widget(self):
        self.label_show_type_label.place(x=30, y=80, width=200, height=50)
        self.draw_type_exchange.place(x=15, y=0)
        self.draw_type_random.place(x=105, y=0)
        self.label_show_gift_label.place(x=370, y=80, width=200, height=50)
        self.gift_type_tianxuan.place(x=15, y=0)
        self.gift_type_safufu.place(x=105, y=0)
        self.start_button.place(x=250, y=88, width=100, height=50)
        self.remain_message.place(x=75, y=150, width=200, height=250)
        self.result_message.place(x=375, y=150, width=175, height=350)
        # self.remain_pr_label.place(x=105, y=420, width=200, height=30)

    def start_button_awake(self):
        self.start_button.config(state=ACTIVE)
        self.remain_message_var.set(self.get_ready_message())

    # 根据抽奖方式选择是否隐藏奖品种类
    def hidden_or_show_gift_type(self):
        value = self.draw_type_button_var.get()
        self.start_button.config(state=ACTIVE)
        if value == 1:
            self.gift_type_button_var.set(0)
            self.remain_message_var.set(self.get_ready_message())
            self.gift_type_tianxuan.config(state=DISABLED)
            self.gift_type_safufu.config(state=DISABLED)
        else:
            self.gift_type_button_var.set(1)
            self.remain_message_var.set(self.get_ready_message())
            self.gift_type_tianxuan.config(state=NORMAL)
            self.gift_type_safufu.config(state=NORMAL)

    # 获取抽奖方式
    def get_draw_type(self):
        if self.start_button["text"] == "开始":
            self.running_flag = True
            if isinstance(self.qunyou_list, list):
                self.start_button.config(text="Bingo")
                if self.draw_type_button_var.get() == 1:
                    mode = 'o2o'
                else:
                    mode = 'random'
                self.thread_it(self.start_draw(mode))
        else:
            self.running_flag = False
            self.start_button.config(text="开始")

    def create_new_window(self):
        width = 215
        height = 100
        left = (self.root.winfo_screenwidth() - width) / 2
        top = (self.root.winfo_screenheight() - height) / 2
        self.new_window = Toplevel(self.root)
        self.new_window.overrideredirect(1)
        self.new_window.geometry("%dx%d+%d+%d" % (width, height, left, top))
        self.new_window.resizable(0, 0)
        label = Label(self.new_window,
                      text="没有惹",
                      font=('Arial', 20, 'bold'))
        label.place(relx=0.35, rely=0.15)
        button = Button(self.new_window,
                        text="好耶",
                        command=self.close_window)
        button.place(relx=0.35, rely=0.5)

    def close_window(self):
        self.new_window.withdraw()
        self.draw_type_random.config(state=ACTIVE)
        self.draw_type_exchange.config(state=ACTIVE)
        self.gift_type_tianxuan.config(state=ACTIVE)
        self.gift_type_safufu.config(state=ACTIVE)

    def is_gift_empty(self):
        if self.qunyou_gift_remain == 0 and self.draw_type_button_var.get() == 1:
            return True
        if self.tianxuan_gift_remain == 0 and self.gift_type_button_var.get() == 1:
            return True
        if self.safufu_gift_remain == 0 and self.gift_type_button_var.get() == 2:
            return True
        return False

    def get_result_message(self):
        return "\n".join(self.result_message_list)

    def get_ready_message(self):
        if self.draw_type_button_var.get() == 1:
            if self.get_ready_gift_index == len(self.qunyou_list):
                return "\n".join(self.qunyou_list + ["Finished."])
            return "\n".join(i + "".join(j) for i, j in self.ready_message_dict.items())
        else:
            if self.gift_type_button_var.get() == 2:
                return "\n".join(self.safufu_temp)
            return "\n".join(self.qunyou_list)

    def start_draw(self, mode: str):
        if self.is_gift_empty():
            self.create_new_window()
            self.start_button.config(text="开始")
            self.start_button.config(state=DISABLED)
            self.draw_type_random.config(state=DISABLED)
            self.draw_type_exchange.config(state=DISABLED)
            self.gift_type_tianxuan.config(state=DISABLED)
            self.gift_type_safufu.config(state=DISABLED)
            return

        if mode == 'o2o':
            temp = self.qunyou_list * 15
            result = set(self.qunyou_list)
            result.difference_update(self.get_gift_list)
            while True:
                if self.running_flag:
                    self.draw_type_random.config(state=DISABLED)
                    random_choice = self.function_random(temp)
                    self.label_show_name_var.set(random_choice)
                    self.label_show_name_adjust(random_choice)
                    time.sleep(self.time_span)
                else:
                    random_choice = self.function_random(list(result))
                    self.label_show_name_var.set(random_choice)
                    self.label_show_name_adjust(random_choice)
                    end_name = self.label_show_name_var.get()
                    result_text = self.qunyou_list[self.get_ready_gift_index] + " ----> " + str(end_name)
                    self.get_gift_list.popleft()
                    self.get_gift_list.append(end_name)
                    self.ready_message_dict[end_name].append(self.STATE2)
                    self.ready_message_dict[self.qunyou_temp[self.get_ready_gift_index]].pop(0)
                    self.qunyou_gift_remain -= 1
                    self.get_ready_gift_index += 1
                    self.ready_message_dict[self.qunyou_temp[self.get_ready_gift_index]].insert(0, self.STATE1)
                    self.result_message_list[self.result_gift_index] = result_text
                    self.get_gift_list.appendleft(self.qunyou_temp[self.get_ready_gift_index])
                    self.result_gift_index += 1
                    # self.gift_remain -= 1
                    self.result_message_var.set(self.get_result_message())
                    self.remain_message_var.set(self.get_ready_message())
                    self.draw_type_random.config(state=ACTIVE)
                    break
        elif mode == 'random':
            temp = self.qunyou_list * 15
            while True:
                if self.running_flag:
                    self.draw_type_exchange.config(state=DISABLED)
                    self.gift_type_tianxuan.config(state=DISABLED)
                    self.gift_type_safufu.config(state=DISABLED)
                    random_choice = self.function_random(temp)
                    self.label_show_name_var.set(random_choice)
                    self.label_show_name_adjust(random_choice)
                    time.sleep(self.time_span)
                else:
                    random_choice = self.function_random(temp)
                    self.label_show_name_var.set(random_choice)
                    self.label_show_name_adjust(random_choice)
                    end_name = self.label_show_name_var.get()
                    gift_type = self.gift_type_button_var.get()
                    result_text = ""
                    if gift_type == 1:
                        result_text = "天选之人" + str(self.tianxuan_count - self.tianxuan_gift_remain + 1) + " ----> " + str(end_name)
                        self.tianxuan_gift_remain -= 1
                    elif gift_type == 2:
                        result = set(self.qunyou_list)
                        result.remove(self.last_year_safufu)
                        random_choice = self.function_random(list(result))
                        self.label_show_name_var.set(random_choice)
                        self.label_show_name_adjust(random_choice)
                        end_name = self.label_show_name_var.get()
                        result_text = "Safufu" + " ----> " + str(end_name)
                        self.safufu_gift_remain -= 1
                    self.result_message_list[self.result_gift_index] = result_text
                    self.result_gift_index += 1
                    self.result_message_var.set(self.get_result_message())
                    self.draw_type_exchange.config(state=ACTIVE)
                    self.gift_type_tianxuan.config(state=ACTIVE)
                    self.gift_type_safufu.config(state=ACTIVE)
                    # self.gift_remain -= 1
                    break

    def thread_it(self, func, *args):
        t = threading.Thread(target=func, args=args)
        t.setDaemon(True)
        t.start()

    def label_show_name_adjust(self, name: str):
        if len(name) == 2:
            self.label_show_name.place(x=245, y=20)
        else:
            self.label_show_name.place(x=220, y=20)

class TipWindow:
    def __init__(self):
        self.window = Tk()
        width = 215
        height = 100
        left = (self.window.winfo_screenwidth() - width) / 2
        top = (self.window.winfo_screenheight() - height) / 2
        self.window.geometry("%dx%d+%d+%d" % (width, height, left, top))
        self.window.resizable(0, 0)
        self.window.title("没有惹")
        button = Button(self.window,
                        text="好耶",
                        command=self.close_window)
        button.place(relx=0.35, rely=0.5)
        self.window.mainloop()

    def close_window(self):
        self.window.destroy()



class ErrorWindow:
    def __init__(self):
        self.window = Tk()
        self.window.title('文件不存在')
        width = 245
        height = 100
        left = (self.window.winfo_screenwidth() - width) / 2
        top = (self.window.winfo_screenheight() - height) / 2
        self.window.geometry("%dx%d+%d+%d" % (width, height, left, top))
        self.window.resizable(0, 0)
        label = Label(self.window,
                      text="找不到文件 peopleList.xlsx",
                      font=('Arial', 15, 'bold'))
        label.place(relx=0.1, rely=0.15)
        button = Button(self.window,
                        text="关闭",
                        command=self.close_window)
        button.place(relx=0.35, rely=0.5)
        self.window.mainloop()

    def close_window(self):
        self.window.destroy()


if __name__ == '__main__':
    try:
        f = open('./peopleList.xlsx')
        f.close()
        app = App()
    except FileNotFoundError:
        window = ErrorWindow()
