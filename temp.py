from tkinter import *
import random
import openpyxl

is_run = False
# def insert_point():
#     var = input.get()
#     return var

def randomResult(people_list: list):
    global is_run
    result = random.choice(people_list)
    label_var.set(result)
    if is_run:
        window.after(100, randomResult, people_list)

def finalResult(people_list: list):
    window.update()
    global is_run
    is_run = False
    temp = label_var.get()
    print(temp)
    result = random.choice(people_list)
    window.update()
    message_var.set(result)

def randomRun(people_list: list):
    global is_run
    if is_run:
        return
    is_run = True
    randomResult(people_list)
    window.update()

def getPeopleList():
    workbook = openpyxl.load_workbook("./peopleList.xlsx", data_only=True)
    list_a = []
    sheet = workbook[workbook.sheetnames[0]]
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == "":
            continue
        else:
            list_a.append(sheet.cell(row=row, column=2).value)
    return list_a


if __name__ == '__main__':
    window = Tk()
    window.geometry('600x450')
    window.resizable(0, 0)
    window.title('火星人专用抽奖 v1.0')

    # 该变量无需改变，每次传入后通过 deepcopy 处理间接参数
    qunyou_list = getPeopleList()
    print(qunyou_list)
    temp_result = []

    label_var = StringVar()
    message_var = StringVar()
    message_var.set('aabb')

    temp_resultLabel = Label(textvariable=label_var)
    temp_resultLabel.place(anchor=NW, x=150, y=100)

    result_message = Message(textvariable=message_var, bg='#ffffff', relief=SUNKEN)
    result_message.place(height=350, width=150, x=400, y=50)

    startButton = Button(text="开始", command=lambda: randomRun(people_list=qunyou_list))
    confirmButton = Button(text="停止", command=lambda: finalResult(people_list=qunyou_list))
    startButton.place(anchor=NW, x=200, y=180)
    confirmButton.place(anchor=NW, x=260, y=180)

    window.mainloop()